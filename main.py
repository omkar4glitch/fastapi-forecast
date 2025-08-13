from fastapi import FastAPI
from pydantic import BaseModel
import pandas as pd
from prophet import Prophet
import os
import uuid
import requests
from io import BytesIO
from fastapi.responses import FileResponse
import openpyxl

app = FastAPI()

# Folder to store temporary files
FILES_DIR = "files"
os.makedirs(FILES_DIR, exist_ok=True)

class ForecastRequest(BaseModel):
    file_url: str
    forecast_start: str  # Example: "2025-08-11"
    forecast_end: str    # Example: "2025-08-17"

@app.post("/forecast")
def forecast(request: ForecastRequest):
    try:
        # === 1. Download file from public URL ===
        response = requests.get(request.file_url)
        response.raise_for_status()

        # === 2. Read Excel or CSV ===
        try:
            df = pd.read_excel(BytesIO(response.content))
        except:
            df = pd.read_csv(BytesIO(response.content))

        # Ensure Date column is datetime
        df["Date"] = pd.to_datetime(df["Date"])

        # === 3. Reshape data to long format ===
        df_long = pd.melt(df, id_vars=["Date"], var_name="Store", value_name="Sales")

        # === 4. Forecast parameters from Adalo ===
        forecast_start = pd.to_datetime(request.forecast_start)
        forecast_end = pd.to_datetime(request.forecast_end)
        future_dates = pd.date_range(start=forecast_start, end=forecast_end, freq='D')
        store_forecasts = {}

        # === 5. Loop over each store and forecast with Prophet ===
        for store in df_long["Store"].unique():
            store_df = df_long[df_long["Store"] == store].dropna()

            if len(store_df) < 2:
                continue  # Skip if not enough data

            prophet_df = store_df.rename(columns={"Date": "ds", "Sales": "y"})

            model = Prophet(daily_seasonality=True, yearly_seasonality=True)
            model.fit(prophet_df)

            future = pd.DataFrame({"ds": future_dates})
            forecast_df = model.predict(future)
            store_forecasts[store] = forecast_df[["ds", "yhat"]].rename(columns={"yhat": store})

        # === 6. Combine forecasts into wide format ===
        wide_forecast = pd.DataFrame({"Date": future_dates})
        for store, store_df in store_forecasts.items():
            wide_forecast = wide_forecast.merge(
                store_df.rename(columns={"ds": "Date"}), 
                on="Date", 
                how="left"
            )

        # === 7. Save output as Excel ===
        output_filename = f"forecast_{uuid.uuid4().hex}.xlsx"
        output_path = os.path.join(FILES_DIR, output_filename)
        wide_forecast.to_excel(output_path, index=False)

        # === 8. Apply formatting ===
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Format Date column as short date
        for cell in ws['A'][1:]:  # skip header
            cell.number_format = 'mm-dd-yyyy'

        # Format sales columns as comma style, no decimals
        for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=2):
            for cell in col:
                cell.number_format = '#,##0'

        wb.save(output_path)

        # === 9. Return public link to file ===
        download_url = f"https://{os.environ.get('RAILWAY_STATIC_URL', 'web-production-df285.up.railway.app')}/files/{output_filename}"

        return {
            "message": "Forecast complete",
            "download_url": download_url
        }

    except Exception as e:
        return {"error": str(e)}

# Serve files
@app.get("/files/{filename}")
def get_file(filename: str):
    file_path = os.path.join(FILES_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename)
    return {"error": "File not found"}
